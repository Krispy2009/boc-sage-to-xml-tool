import os
import re
import pdb
import uuid
import datetime
import xml.etree.ElementTree as ET

from pydoc import doc
from turtle import pd

from openpyxl import load_workbook
from constants import BOC_BIC


def build_grp_hdr(ccti, number_of_txns=0, total=0):
    grp_hdr = ET.SubElement(ccti, "GrpHdr")
    msg_id = ET.SubElement(grp_hdr, "MsgId")
    msg_id.text = str(uuid.uuid4().hex)

    # TODO: Correct format
    dt = ET.SubElement(grp_hdr, "CreDtTm")
    dt.text = str(datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S"))

    txns = ET.SubElement(grp_hdr, "NbOfTxs")
    txns.text = str(number_of_txns)

    # Control Sum has 2 decimal places
    control_sum = ET.SubElement(grp_hdr, "CtrlSum")
    control_sum.text = f"{total:.2f}"

    init_party = ET.SubElement(grp_hdr, "InitgPty")
    nm = ET.SubElement(init_party, "Nm")
    nm.text = f"{os.environ['COMPANY_NAME']}"


def build_pmts(ccti, transaction):
    if transaction.get("IBAN") is None:
        # A critical piece is missing don't try to build pmt
        return

    pmt_info = ET.SubElement(ccti, "PmtInf")
    pmt_id = ET.SubElement(pmt_info, "PmtInfId")
    pmt_id.text = str(uuid.uuid4().hex)

    # Value is always TRF which means Credit Transfer
    pmt_mtd = ET.SubElement(pmt_info, "PmtMtd")
    pmt_mtd.text = "TRF"

    pmt_tp_inf = ET.SubElement(pmt_info, "PmtTpInf")
    service_level = ET.SubElement(pmt_tp_inf, "SvcLvl")
    code = ET.SubElement(service_level, "Cd")
    if transaction.get("BIC") == BOC_BIC:
        code.text = "TBOC"
    else:
        code.text = "SEPA"

    req_ex_date = ET.SubElement(pmt_info, "ReqdExctnDt")
    req_ex_date.text = datetime.datetime.now().strftime("%Y-%m-%d")

    # The only required fields is IBAN, BIC
    # Charges are by default Shared in SEPA and OUR in TBOC transfers
    # TODO: Ask whether this is ok
    debtor = ET.SubElement(pmt_info, "Dbtr")

    debtor_acct = ET.SubElement(pmt_info, "DbtrAcct")
    debtor_id = ET.SubElement(debtor_acct, "Id")
    iban = ET.SubElement(debtor_id, "IBAN")
    iban.text = str(os.environ["COMPANY_IBAN"])

    debtor_agent = ET.SubElement(pmt_info, "DbtrAgt")
    fin_inst = ET.SubElement(debtor_agent, "FinInstnId")
    bic = ET.SubElement(fin_inst, "BIC")
    bic.text = BOC_BIC

    # Not Sure this is needed
    # TODO: Ask about charges (like above)
    chrg_bearer = ET.SubElement(pmt_info, "ChrgBr")
    if transaction.get("BIC") == BOC_BIC:
        chrg_bearer.text = "DEBT"
    else:
        chrg_bearer.text = "SHAR"

    txn_info = ET.SubElement(pmt_info, "CdtTrfTxInf")

    tx_pmt_id = ET.SubElement(txn_info, "PmtId")
    e2e = ET.SubElement(tx_pmt_id, "EndToEndId")
    e2e.text = str(uuid.uuid4().hex)

    # TODO: CHECK THIS - is it always P2 and P3?
    amt = ET.SubElement(txn_info, "Amt")
    inst_amt = ET.SubElement(amt, "InstdAmt", attrib={"Ccy": "EUR"})
    inst_amt.text = str(transaction["Period 2"] + transaction["Period 3"])

    creditor = ET.SubElement(txn_info, "Cdtr")
    cr_nm = ET.SubElement(creditor, "Nm")
    cr_nm.text = transaction["name"][:34]

    creditor_acct = ET.SubElement(txn_info, "CdtrAcct")
    creditor_id = ET.SubElement(creditor_acct, "Id")
    iban = ET.SubElement(creditor_id, "IBAN")
    iban.text = str(transaction["IBAN"])

    txn_info = add_remit_info(txn_info, transaction)


def add_remit_info(txn_info, transaction):
    # A message that will appear on the creditor's statement
    if transaction.get("notes"):
        info = ET.SubElement(txn_info, "RmtInf")
        msg = ET.SubElement(info, "Ustrd")
        msg.text = transaction["notes"]
    return txn_info


def build_xml(transactions):
    document = ET.Element(
        "Document",
        attrib={
            "xmlns:NS1": "http://www.w3.org/2001/XMLSchema-instance",
            "xmlns": "urn:iso:std:iso:20022:tech:xsd:pain.001.001.03",
            "NS1:schemaLocation": "urn:iso:std:iso:20022:tech:xsd:pain.001.001.03 pain.001.001.03.xsd",
        },
    )
    ccti = ET.SubElement(document, "CstmrCdtTrfInitn")

    # TODO: Check if this is right
    total = sum([i["total_p2_p3"] for i in transactions])

    # Add Group Header
    build_grp_hdr(ccti, number_of_txns=len(transactions), total=total)

    # One PMT Info element per transaction
    for transaction in transactions:
        build_pmts(ccti, transaction)

    filename = f"{datetime.datetime.now().strftime('%Y%m%d_%H%M%S_ACMA.xml')}"
    with open(filename, "wb") as f:
        f.write(ET.tostring(document))


def parse_sage_report(filename):
    transactions = []
    wb = load_workbook(filename)
    ws = wb.active

    # Keep unmerging until there are no merged cells
    while len(ws.merged_cells.ranges):
        ws = unmerge_rows(ws)

    for xl_row in ws:
        if xl_row[0].row < 8:
            continue
        row_to_append = xl_row_to_dict(xl_row)
        if row_to_append.get("IBAN") is not None:
            transactions.append(row_to_append)

    return transactions


def do_calcs(txns):
    for tx in txns:
        total_p1_p2 = tx["Period 1"] + tx["Period 2"]
        tx["total_p1_p2"] = total_p1_p2

        total_p2_p3 = tx["Period 2"] + tx["Period 3"]
        tx["total_p2_p3"] = total_p2_p3

        total_p1_p2_p3 = tx["Period 1"] + tx["Period 2"] + tx["Period 3"]
        tx["total_p1_p2_p3"] = total_p1_p2_p3

        total_all = tx["Period 1"] + tx["Period 2"] + tx["Period 3"] + tx["Older"]
        tx["total_all"] = total_all

    return txns


def xl_row_to_dict(xl_row):
    row = {}
    row_map = {
        "B": "name",
        "F": "BIC",
        "H": "IBAN",
        "K": "Currency",
        "L": "Turnover",
        "M": "Balance",
        "N": "Current",
        "O": "Period 1",
        "P": "Period 2",
        "Q": "Period 3",
        "R": "Older",
    }
    for col in xl_row:

        if col.column_letter in row_map:
            row_title = row_map[col.column_letter]
            value = col.value
            if row_title == "IBAN":
                value = remove_spaces(col.value)
            elif row_title == "name":
                value = remove_special_chars(col.value)
            row[row_title] = value
    return row


def unmerge_rows(ws):

    for merged in ws.merged_cells:
        ws.unmerge_cells(str(merged))
    return ws


def remove_spaces(iban):
    if iban:
        return iban.replace(" ", "")


def remove_special_chars(name):
    # acceptable_special_chars = ["/", "-", "?", ":", "(", ")", ".", ",", "‘", "+", " "]
    RE = re.compile(r"[!@&*\]\[{};_><^%$#]+")
    if name:
        for match in RE.findall(name):
            name = name.replace(match, "")
    return name


if __name__ == "__main__":
    transactions = parse_sage_report("sample.xlsx")
    transactions = do_calcs(transactions)
    build_xml(transactions)
