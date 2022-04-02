import re
from openpyxl import load_workbook


def unmerge_rows(ws):

    for merged in ws.merged_cells:
        ws.unmerge_cells(str(merged))
    return ws


def remove_spaces(iban):
    if iban:
        return iban.replace(" ", "")


def remove_special_chars(name):
    # acceptable_special_chars = ["/", "-", "?", ":", "(", ")", ".", ",", "‘", "+", " "]
    RE = re.compile(r"[!@&*\]\[{};_><^%$#]+")
    if name:
        for match in RE.findall(name):
            name = name.replace(match, "")
    return name


def xl_row_to_dict(xl_row):
    row = {}
    row_map = {
        "B": "name",
        "F": "BIC",
        "H": "IBAN",
        "K": "Currency",
        "L": "Turnover",
        "M": "Balance",
        "N": "Current",
        "O": "Period 1",
        "P": "Period 2",
        "Q": "Period 3",
        "R": "Older",
    }
    for col in xl_row:

        if col.column_letter in row_map:
            row_title = row_map[col.column_letter]
            value = col.value
            if row_title == "IBAN":
                value = remove_spaces(col.value)
            elif row_title == "name":
                value = remove_special_chars(col.value)
            row[row_title] = value
    return row


def parse_sage_report(filename):
    transactions = []
    wb = load_workbook(filename)
    ws = wb.active

    # Keep unmerging until there are no merged cells
    while len(ws.merged_cells.ranges):
        ws = unmerge_rows(ws)

    for xl_row in ws:
        if xl_row[0].row < 8:
            continue
        row_to_append = xl_row_to_dict(xl_row)
        if row_to_append.get("IBAN") is not None:
            transactions.append(row_to_append)

    return transactions


def do_calcs(txns):
    for tx in txns:
        total_p1_p2 = tx["Period 1"] + tx["Period 2"]
        tx["total_p1_p2"] = total_p1_p2

        total_p2_p3 = tx["Period 2"] + tx["Period 3"]
        tx["total_p2_p3"] = total_p2_p3

        total_p1_p2_p3 = tx["Period 1"] + tx["Period 2"] + tx["Period 3"]
        tx["total_p1_p2_p3"] = total_p1_p2_p3

        total_all = tx["Period 1"] + tx["Period 2"] + tx["Period 3"] + tx["Older"]
        tx["total_all"] = total_all

    return txns
